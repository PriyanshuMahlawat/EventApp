from django.shortcuts import render
import requests
import os
from django.conf import settings
from django.shortcuts import get_object_or_404
import openpyxl
import pandas as pd
from django.utils import timezone
from django.core.files.base import ContentFile
import io
from datetime import datetime
import pytz
from django.http import HttpResponse,FileResponse
from openpyxl import Workbook
import cloudinary.uploader
from django.contrib.auth.models import User
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.csrf import csrf_exempt
import json
from django.urls import reverse
from .forms import EventForm
from .models import Event,Notifications,CurrentEvent,slots,completedEvents,RealTable,FinalSlotsTable,tablemodifications
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_protect
from allauth.socialaccount.models import SocialAccount
from django.db import IntegrityError

from rest_framework import generics
from rest_framework.views import APIView

from django.db import IntegrityError
from rest_framework.response import Response
from .serializers import (
    EventSerializer,
    EventDetailSerializer,
    NotiSerializer,
    HostedEventsSerializer,
    AddingMembersSerializer,
    JoinedEventsSerializer,
    ManageEventSerializer,
    MembersListSerializer,
    CurrentEventSerializer,
    slotsSerializer,
    tableSlotSerializer,
    CompletedEventsSerializer,
    CurrentTableSerializer,
    RealTableSerializer,
    CompletedEventGetSerializer,
    FinalSlotsSerializer,
    tablemodificationsSerializer
)







def completedpage(request):
    if 'event_id' in request.COOKIES:
        event_id = request.COOKIES.get('event_id')
    else:
        event_id = -1
    
    user= request.user
    if user.is_authenticated:
        username = user.username
    else:
        username = None

    
    if Event.objects.filter(id=event_id).exists() and completedEvents.objects.filter(event_id = event_id).exists():
        Event.objects.filter(id=event_id).delete()  

    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture') 
    return render(request,"app/completed.html",{"username":username,"logged_in":logged_in,"image_url":image_url})


def suggestpage(request):
    context= {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration

    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    context["logged_in"]=logged_in
    context["image_url"]=image_url
    return render(request,"app/suggested.html",context)

def create_slots():
    
    events = Event.objects.all()
    for e in events:
        event_id = e.id
        memberstr = e.members
        if memberstr:
            usernames = memberstr.split(",")
            for username in usernames:
                print(username)
                print(event_id)
                try:
                    user = User.objects.get(username=username)
                    slot, created = slots.objects.get_or_create(user=user, event_id=event_id)
                    if created:
                        print(f"Slot for user {username} and event {event_id} created.")
                    else:
                        print(f"Slot for user {username} and event {event_id} already exists.")
                except User.DoesNotExist:
                    print(f"User with username {username} does not exist.")
                except IntegrityError as e:
                    print(f"IntegrityError: {e} for user {username} and event {event_id}.")




def index(request):
    logged_in = ""
    image_url = None
    
    if request.user.is_authenticated:
        logged_in ="1"
        person = request.user.username
        email = request.user.email
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
   
    else:
        person = None
        email = None
    
    return render(request,"app/userpage.html",{"person":person,"email":email,"logged_in":logged_in,"image_url":image_url})

def myeventtab(request):
    logged_in = False
    image_url = None
    
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    return render(request,"app/myevent.html",{"logged_in":logged_in,"image_url":image_url})

def about(request):
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    return render(request,'app/about.html',{"logged_in":logged_in,"image_url":image_url})


def admin(request):
    context = {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    username = request.user.username
    context['username'] = username
    return render(request,"app/admin.html",context)


def editEvent(request):
    context = {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
        context['image_url'] = image_url
        context['logged_in'] = logged_in
    return render(request,"app/editevent.html",context)

def DetailEvent(request):
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    return render(request,"app/eventdetail.html",{"logged_in":logged_in,"image_url":image_url})

def noti(request):
    context = {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')

    person = request.user.username
    host = False
    event_id = request.COOKIES.get('event_id')
    
    if(request.user.username == Event.objects.get(id = event_id).host.username):
        host = True
    context['image_url'] = image_url
    context['person'] = person
    context['host'] = host
    context['logged_in'] = logged_in
    return render(request,"app/notifications.html",context)

def slotspg(request):
    context = {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    host = False
    event_id = request.COOKIES.get('event_id')
    name=request.user.username
    if(request.user.username == Event.objects.get(id = event_id).host.username):
        host = True
    context['image_url'] = image_url
    context['host'] = host
    context['logged_in'] = logged_in
    return render(request,"app/slots.html",context)







def roles(request):
    context = {}
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    context['image_url'] = image_url
    
    context['logged_in'] = logged_in
    return render(request,"app/roles.html",context)



@csrf_exempt
def manage(request):
    logged_in = False
    image_url = None
    member = False
    Nfinalised = True
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    host = False
    event_id = request.COOKIES.get('event_id')
    event = Event.objects.get(id=event_id)
    
    if FinalSlotsTable.objects.filter(Event_id =event_id).exists():
        Nfinalised = False
    try:
        if event.members:
            if request.user.username in event.members.split(','):
                member = True
        if request.user.username == event.host.username:
            host = True
    except Event.DoesNotExist:
        print("Event not found")
        return render(request, "app/manage.html", {"error": "Event not found"})

    person = request.user.username
    print(Nfinalised)
    context = {
        "member":member,
        "person": person,
        "live": False,
        "1v": False,
        "2v": False,
        "3v": False,
        "4v": False,
        "host":False,
        "Nfinalised":Nfinalised,
    }
    
    if host:
        
        context["1v"] = True
        context["2v"] = True
        context["3v"] = True
        context["4v"] = True
        context["host"] = True
    event = Event.objects.get(id=event_id)
    timezone1 = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone1)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration
    print(timer_start)
    print(datetime.now())
    if timezone.now() >= timer_start:
        context['live'] = True
    
        
    user = request.user
    
    
    if user.groups.filter(name=f"Add Members{event_id}").exists():
        context["1v"] = True
    if user.groups.filter(name=f"Allow Noti{event_id}").exists():
        context["2v"] = True
    if user.groups.filter(name=f"Remove Members{event_id}").exists():
        context["3v"] = True
    if user.groups.filter(name=f"Edit Event{event_id}").exists():
        context["4v"] = True

    context["logged_in"] = logged_in
    context["image_url"] = image_url
    
    
    

    return render(request, "app/manage.html", context)


@login_required
def slotnoti(request):
    context ={}
    event_id = request.COOKIES.get('event_id')
    
    try:
            finalmodel = FinalSlotsTable.objects.get(Event__id=event_id)
            link = finalmodel.download_url
            context['link'] = link
    except:
        context['link']=None


    event = get_object_or_404(Event,id=event_id)
    
    
    timezone = pytz.timezone('Asia/Kolkata')
    timer_start = event.event_time
    duration = str(event.event_duration.total_seconds())
    timer_start = timer_start.astimezone(timezone)
    timeStart = timer_start.isoformat()
    context["timer_start"] = timeStart
    context['duration']= duration

    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    context["logged_in"]=logged_in
    context["image_url"]=image_url
    return render(request,"app/slotnoti.html",context)



@login_required
def hostpg(request):
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    success = request.session.pop('success', False)
    error = request.session.pop('error', [])
    event_name = request.session.pop('event_name', None)
    
    if request.method == "POST":
        
        event_form = EventForm(request.POST,request.FILES)
        if event_form.is_valid():
            event_form.save(user=request.user)
            event_name = event_form.cleaned_data['Event_name']
            request.session['success'] = True
            request.session['event_name'] = event_name
        else:
            request.session['success'] = False
            request.session['error'] = list(event_form.errors.values())
        return redirect('popup')
    else:
        event_form = EventForm()
    
    return render(request, "app/hostpage.html", {
        "event_form": event_form,
        "success": success,
        "error": error,
        "event_name": event_name,
        "logged_in":logged_in,
        "image_url":image_url,
    })


@csrf_protect
def popup_view(request):
    logged_in = False
    image_url = None
    if request.user.is_authenticated:
        logged_in =True
        
       
        social_account = SocialAccount.objects.get(user=request.user)
        data = social_account.extra_data
        image_url = data.get('picture')
    context = {}
    context['success'] =request.session.get('success',False)
    context['event_name'] =request.session.get('event_name',None)
    context['error'] =request.session.get('error',[])
    context['logged_in'] = logged_in
    context['image_url'] = image_url
    return render(request,"app/popup.html",context)





class SlotsDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = slots.objects.all()
    serializer_class = slotsSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        """Fetch the slot for the current user and event, or None if not found."""
        try:
            return slots.objects.get(user=self.request.user, event_id=self.kwargs['Event_id'])
        except slots.DoesNotExist:
            return None

    def update(self, request, *args, **kwargs):
        event_id = kwargs.get('Event_id')
        # Try to retrieve the slot
        instance = self.get_object()

        if instance is None:
            # Slot does not exist, so create a new one
            instance = slots.objects.create(
                user=self.request.user,
                event_id=event_id,
                slots=request.data.get('slots', {})
            )
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        else:
            # Update existing slot
            instance.slots = request.data.get('slots', {})
            instance.save()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)

        



class EventSerializerAPIView(generics.ListAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer


class EventDetailSerializerAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Event.objects.all()
    serializer_class = EventDetailSerializer
    lookup_field = 'pk'


class NotiSerializerAPIView(generics.ListCreateAPIView):
    queryset = Notifications.objects.all()
    serializer_class = NotiSerializer

    


class NotiDestroyUpdateView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Notifications.objects.all()
    serializer_class = NotiSerializer 

class HostedEventSerializerAPIView(generics.ListCreateAPIView):
    queryset = Event.objects.all()
    serializer_class = HostedEventsSerializer

    def get_serializer_context(self):
        
        context = super().get_serializer_context()
        context.update({
            'request': self.request,  
        })
        return context
    

class HostedEventsUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Event.objects.all()
    serializer_class = HostedEventsSerializer

    def get_serializer_context(self):
        context =  super().get_serializer_context() 
        context.update({
            "request":self.request,
        }) 
        return context  
    
class membersRetrieveAPIView(generics.RetrieveAPIView):
    queryset = Event.objects.all()
    serializer_class = MembersListSerializer
    


class membersUpdateAPIView(generics.UpdateAPIView):
    queryset = Event.objects.all()
    serializer_class = MembersListSerializer
    


class AddmembersAPIView(generics.UpdateAPIView):
    queryset = Event.objects.all()
    serializer_class = AddingMembersSerializer

    
    def update(self,request,*args,**kwargs):
        instance  =self.get_object()
        new_member = request.data.get('members','')
        existing_members = instance.members.split(',') if instance.members else []
        if new_member not in existing_members:
            
            if instance.members:
                instance.members += f",{new_member}"
            else:
                instance.members = new_member

        instance.save()
        serializer = self.get_serializer(instance)
        create_slots()
        return Response(serializer.data)      


class JoinedAPIView(generics.ListAPIView):
    queryset = Event.objects.all()
    serializer_class = JoinedEventsSerializer

    def get_serializer_context(self):
        context =  super().get_serializer_context() 
        context.update({
            "request":self.request,
        }) 
        return context 

class LeaveEventAPIView(APIView):
    def delete(self, request, event_id):
        try:
            event = Event.objects.get(id=event_id)
            name = request.user.username
            if name in event.members:
                members_list = event.members.split(',')
                new_list = ",".join(member for member in members_list if member != name)
                event.members = new_list
                event.save()
                return Response({"message": "Successfully left the event."})
            else:
                return Response({"error": "Something is wrong. You are not in the members list."})
        except Event.DoesNotExist:
            return Response({"error": "Event not found."}, status=404)






class ManageEventAPI(generics.RetrieveUpdateAPIView):
    queryset = Event.objects.all()
    serializer_class = ManageEventSerializer         



class CurrentEventAPIView(APIView):
    

    def get(self, request):
        current_event, created = CurrentEvent.objects.get_or_create(user=request.user)
        serializer = CurrentEventSerializer(current_event)
        return Response(serializer.data)

    def post(self, request):
        
        current_event, created = CurrentEvent.objects.get_or_create(user=request.user)
        
            
            
        serializer = CurrentEventSerializer(current_event, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
    


class tableSlotAPIView(APIView):

    def get(self,request,**kwargs):
        event_id = kwargs.get('pk')
        event = Event.objects.get_or_create(id =event_id)
        serializer = tableSlotSerializer(event,context={'id':event_id})
        return Response(serializer.data)


class TableModificationsAPIView(APIView):

    def get(self,request,**kwargs):
        event_id = kwargs.get('pk')
        model = tablemodifications.objects.get(Event__id = event_id)
        serializer = tablemodificationsSerializer(model)

        return Response(serializer.data['table'])
    
    def fix(self,opTable,startTime,endTime,name,roomNew):
        room_slots = opTable['slot'][roomNew]
        i=0
        inserted =False
        while i < len(room_slots):
                print('went in')
                current_slot = room_slots[i]
                current_start_time, current_end_time = list(current_slot.values())[0]
                current_username = list(current_slot.keys())[0]
                
                
                
                if current_username != 'null' and name != 'null':
                    new_name = f"{current_username}_{name}"
                elif name != 'null':
                    new_name = name  
            
                if (startTime) == (current_start_time) and (endTime) < (current_end_time):
                    print("check1")
                    room_slots[i] = {new_name: (startTime, endTime)}
                    room_slots.insert(i + 1, {current_username: (endTime, current_end_time)})
                    inserted = True
                    break

                elif (startTime) == (current_start_time) and (endTime) > (current_end_time):
                    print("check2")
                    room_slots[i] = {new_name: (current_start_time, current_end_time)}
                    room_slots.insert(i + 1, {f"{name}": (current_end_time, endTime)})
                    inserted = True
                    break

                elif (startTime) == (current_start_time) and (endTime) == (current_end_time):
                    print("check3")
                    room_slots[i] = {new_name: (startTime, endTime)}
                    inserted = True
                    break

                elif (startTime) > (current_start_time) and (startTime) < (current_end_time):
                    if (endTime) < (current_end_time):
                        print("check4")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        room_slots.insert(i + 1, {new_name: (startTime, endTime)})
                        room_slots.insert(i + 2, {current_username: (endTime, current_end_time)})
                        inserted = True
                        break

                    elif (endTime) == (current_end_time):
                        print("check5")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        room_slots.insert(i + 1, {new_name: (startTime, endTime)})
                        inserted = True
                        break

                    elif (endTime) > (current_end_time):
                        print("check6")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        room_slots.insert(i + 1, {new_name: (startTime, current_end_time)})
                        room_slots.insert(i + 2, {name: (current_end_time, endTime)})
                        inserted = True
                        break

                elif (startTime) < (current_start_time) and (endTime) < (current_end_time):
                    print("check7")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    room_slots.insert(i + 1, {new_name: (current_start_time, endTime)})
                    room_slots.insert(i + 2, {current_username: (endTime, current_end_time)})
                    inserted = True
                    break

                elif (startTime) < (current_start_time) and (endTime) > (current_end_time):
                    print("check8")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    room_slots.insert(i + 1, {new_name: (current_start_time, current_end_time)})
                    room_slots.insert(i + 2, {name: (current_end_time, endTime)})
                    inserted = True
                    break

                elif (startTime) < (current_start_time) and (endTime) == (current_end_time):
                    print("check9")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    room_slots.insert(i + 1, {new_name: (current_start_time, current_end_time)})
                    inserted = True
                    break

                i += 1

           
        if not inserted:
                room_slots.append({name: (startTime, endTime)})

        
        opTable['slot'][roomNew] = room_slots

        return opTable
    

    def clean(self, table, room):
        
            room_slots = table['slot'][room]
            cleaned_slots = []
            
            for slot in room_slots:
            
                for key in slot:
                    if key is not None:
                        usernamesArr = key.split('_')
                        uniqueArr = list(set(usernamesArr))
                        
                        
                        cleaned_slot = {('_'.join(uniqueArr)): slot[key]}
                        cleaned_slots.append(cleaned_slot)
                    else:
                        cleaned_slot = {None: slot[key]}
                        cleaned_slots.append(cleaned_slot)  
            
            table['slot'][room] = cleaned_slots
        
            return table


    def patch(self,request,**kwargs):
        event_id = kwargs.get('pk')
        
        changes = request.data.get('changes')
        modeltable = tablemodifications.objects.get(Event__id=event_id)
        Arr = changes
        nameString = Arr[0]
        name = Arr[1]
        roomIni= Arr[2] 
        roomNew = Arr[3]
        startTime = Arr[4] +":00"
        endTime = Arr[5]+":00"
        opTable = modeltable.table 
        _count = nameString.count('_')
        if _count ==0:
            new_name= 'null'
        else:
            new_name = nameString.replace(f"_{name}", "")
        roomsdictIni = opTable['slot'][roomIni] 

        
        
        
        found = False 
        
       
        for slot in roomsdictIni:
            if nameString in slot:
                slot_times = slot[nameString]
                
                if slot_times == [startTime, endTime]:
                    found = True

                    slot[new_name] = slot.pop(nameString) 
                    
                    break
        
        if not found:
            print(f"{nameString} with time slot {startTime} - {endTime} not found in any slot.")
    
    
        opTable['slot'][roomIni] = roomsdictIni
       
            
        room_slots = opTable['slot'][roomNew]
        print(f'room_slots{room_slots}')
        print(len(room_slots))
        i = 0
        inserted = False
        
        
            
        while i < len(room_slots):
                print('went in')
                current_slot = room_slots[i]
                current_start_time, current_end_time = list(current_slot.values())[0]
                current_username = list(current_slot.keys())[0]
                
                
                
                if current_username != 'null' and name != 'null':
                    new_name = f"{current_username}_{name}"
                elif name != 'null':
                    new_name = name  
            
                if (startTime) == (current_start_time) and (endTime) < (current_end_time):
                    print("check1")
                    room_slots[i] = {new_name: (startTime, endTime)}
                    self.fix(opTable,endTime,current_end_time,current_username,roomNew)
                    
                    inserted = True
                    break

                elif (startTime) == (current_start_time) and (endTime) > (current_end_time):
                    print("check2")
                    room_slots[i] = {new_name: (current_start_time, current_end_time)}
                    
                    self.fix(opTable,current_end_time,endTime,name,roomNew)
                    inserted = True
                    break

                elif (startTime) == (current_start_time) and (endTime) == (current_end_time):
                    print("check3")
                    room_slots[i] = {new_name: (startTime, endTime)}
                    inserted = True
                    break

                elif (startTime) > (current_start_time) and (startTime) < (current_end_time):
                    if (endTime) < (current_end_time):
                        print("check4")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        self.fix(opTable,startTime,endTime,new_name,roomNew)
                        
                        self.fix(opTable,endTime,current_end_time,current_username,roomNew)
                        
                        inserted = True
                        break

                    elif (endTime) == (current_end_time):
                        print("check5")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        self.fix(opTable,startTime,endTime,new_name,roomNew)
                        
                        inserted = True
                        break

                    elif (endTime) > (current_end_time):
                        print("check6")
                        room_slots[i] = {current_username: (current_start_time, startTime)}
                        self.fix(opTable,startTime,current_end_time,new_name,roomNew)
                        
                        self.fix(opTable,current_end_time,endTime,name,roomNew)
                        
                        inserted = True
                        break

                elif (startTime) < (current_start_time) and (endTime) < (current_end_time):
                    print("check7")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    self.fix(opTable,current_start_time,endTime,new_name,roomNew)
                    
                    self.fix(opTable,endTime,current_end_time,current_username,roomNew)
                    
                    inserted = True
                    break

                elif (startTime) < (current_start_time) and (endTime) > (current_end_time):
                    print("check8")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    self.fix(opTable,current_start_time,current_end_time,new_name,roomNew)
                    
                    self.fix(opTable,current_end_time,endTime,name,roomNew)
                    
                    inserted = True
                    break

                elif (startTime) < (current_start_time) and (endTime) == (current_end_time):
                    print("check9")
                    room_slots[i] = {name: (startTime, current_start_time)}
                    self.fix(opTable,current_start_time,current_end_time,new_name,roomNew)
                    
                    inserted = True
                    break

                i += 1

           
        if not inserted:
                room_slots.append({name: (startTime, endTime)})

        
        opTable['slot'][roomNew] = room_slots
        opTable = self.clean(opTable,roomNew)

        print(opTable)
        serializer = tablemodificationsSerializer(modeltable, data={'table': opTable}, partial=True)
        if serializer.is_valid():
            serializer.save()
        return Response(serializer.data)


    
    


    def post(self,request,**kwargs):
        
        event_id  =kwargs.get('pk')
        try:
            event = Event.objects.get(id = event_id)
        except:
            return Response({"error":"Event not found"})
        
        try:
       
            table_modification = tablemodifications.objects.get(Event=event)
            serializer = tablemodificationsSerializer(table_modification, data=request.data)  
        except tablemodifications.DoesNotExist:
        
            serializer = tablemodificationsSerializer(data=request.data)
            
        if serializer.is_valid():
            serializer.save(Event=event)
            return Response(serializer.data)
        return Response(serializer.errors)


  
class FinalTableAPIView(generics.CreateAPIView):
    queryset = FinalSlotsTable.objects.all()
    serializer_class = FinalSlotsSerializer

    def create(self, request, *args, **kwargs):
        event_id = kwargs.get('pk')
        
        # Fetch data from API
        endpoint = f"http://localhost:8000/api/tablemodify/{event_id}/"
        response = requests.get(endpoint)
        if response.status_code != 200:
            return Response({"error": "Failed to fetch data from API"}, status=500)
            
        data = response.json()
        print("Fetched data:", data)
        
        try:
            event = Event.objects.get(id=event_id)
            Ename = event.Event_name
            
            # Prepare rows for Excel
            rows = []
            for room_id, slots in data['slot'].items():
                for slot in slots:
                    user = list(slot.keys())[0]
                    times = slot[user]
                    rows.append({
                        "Room": room_id,
                        'Name': user if user != 'null' else 'Empty',
                        'Start Time': times[0],
                        'End Time': times[1],
                    })
                # Add empty rows for visual separation
                rows.append({"Room": '', 'Name': '', 'Start Time': '', 'End Time': ''})
                rows.append({"Room": '', 'Name': '', 'Start Time': '', 'End Time': ''})

            # Create Excel file in memory
            df = pd.DataFrame(rows)
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            excel_buffer.seek(0)
            
            # Generate a unique filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_file_name = f"{Ename}_eventSlots_{timestamp}.xlsx"
            
            # Upload to Cloudinary directly
            cloudinary_response = cloudinary.uploader.upload(
                excel_buffer,
                resource_type="raw",
                public_id=f"excel_files/{excel_file_name}",
                folder="event_slots",
                format="xlsx"
            )
            
            # Create or update FinalSlotsTable instance
            model, created = FinalSlotsTable.objects.get_or_create(Event=event)
            model.excel_file = cloudinary_response['secure_url']
            model.save()
            
            # Return response
            serializer = self.get_serializer(model)
            return Response(serializer.data)
            
        except Event.DoesNotExist:
            return Response({"error": "Event not found"}, status=404)
        except Exception as e:
            print(f"Error processing request: {str(e)}")
            return Response({"error": str(e)}, status=500)

        


        
        



class CompletedEventsSerializerView(generics.CreateAPIView):
    queryset = completedEvents.objects.all()
    serializer_class = CompletedEventsSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        event_id = kwargs.get('event_id')

        # Initialize Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = f'Event_Report-{event_id}'
        
        ws.append(['Username', 'Room', 'Time_Spent', 'Total_Work_Time'])
        row = 2
        for username, user_data in data.items():
            records = user_data.get('records', [])
            n = len(records)
            total_time = sum(record.get('time_spent', 0) for record in records)
            ws.merge_cells(start_row=row, start_column=1, end_row=row + n - 1, end_column=1)
            ws.merge_cells(start_row=row, start_column=4, end_row=row + n - 1, end_column=4)
            ws.cell(row=row, column=1, value=username)
            ws.cell(row=row, column=4, value=total_time)
            for record in records:
                ws.cell(row=row, column=2, value=record['room'])
                ws.cell(row=row, column=3, value=record['time_spent'])
                row += 1

        # Save the workbook to an in-memory buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        try:
            event = Event.objects.get(id=event_id)
            
            # Generate a unique filename and upload to Cloudinary
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"{event.Event_name}_Event_Records_{timestamp}.xlsx"

            cloudinary_response = cloudinary.uploader.upload(
                excel_buffer,
                resource_type="raw",
                public_id=f"excel_files/{file_name}",
                folder="event_records",
                format="xlsx"
            )

            # Create or update CompletedEvents instance with Cloudinary URL
            model, created = completedEvents.objects.update_or_create(
                id=event_id,
                defaults={
                    'event_id': event_id,
                    'host': event.host.username,
                    'Event_name': event.Event_name,
                    'members': event.members,
                    'excel': cloudinary_response['secure_url']
                }
            )

            # Serialize and return response
            serializer = self.get_serializer(model)
            return Response(serializer.data)

        except Event.DoesNotExist:
            return Response({"error": "Event not found"}, status=404)
        except Exception as e:
            print(f"Error processing request: {str(e)}")
            return Response({"error": str(e)}, status=500)
    
    
        



class CompletedEventGetAPIView(APIView):
    def get(self, request, *args, **kwargs):
        response_data = {
            'hosted': [],
            'joined': [],
        }
        
        username = request.user.username
        
        
        for event in completedEvents.objects.all():
            
            if username == event.host and event.excel:
                record = {
                    'Event_name': event.Event_name,
                    'excel_link': event.excel  
                }
                response_data['hosted'].append(record)

            
            if username in event.members.split(',') and event.excel and username != event.host:
                record = {
                    'Event_name': event.Event_name,
                    'excel_link': event.excel 
                }
                response_data['joined'].append(record)
        print(f"data{response_data}")
        return Response(response_data)

            
        
        
            
    
class TableView(APIView):
    table_storage = {}  

    def get(self, request, event_id, *args, **kwargs):
        
        current_table = self.table_storage.get(event_id, {})
        serializer = CurrentTableSerializer({'table': current_table})
        return Response(serializer.data)

    def post(self, request, event_id, *args, **kwargs):
        
        serializer = CurrentTableSerializer(data=request.data)
        if serializer.is_valid():
            self.table_storage[event_id] = serializer.validated_data['table']
            return Response({"message": "Table data updated successfully."})
        return Response(serializer.errors)


class RealTableView(generics.CreateAPIView):
    queryset = RealTable.objects.all()
    serializer_class = RealTableSerializer

    def post(self, request, *args, **kwargs):
        event_id = kwargs.get('event_id')
        username = kwargs.get('username')
        data = request.data
        

        try:
            instance = RealTable.objects.get(event_id=event_id)
        except RealTable.DoesNotExist:
            
            instance = RealTable(event_id=event_id, table={})
        
        table = instance.table or {}

        if username not in table:
            table[username] = {'records': []}
        
        for record in data.get(username, {}).get('records', []):
            table[username]['records'].append(record)
        
        instance.table = table
        instance.save()
        
        serializer = RealTableSerializer(instance)
        return Response(serializer.data)

    def get(self, request, *args, **kwargs):
        event_id = kwargs.get('event_id')
        try:
            table = RealTable.objects.get(event_id=event_id)
            serializer = RealTableSerializer(table)
            return Response(serializer.data)
        except RealTable.DoesNotExist:
            return Response({"error": "RealTable does not exist"}, status=404)

